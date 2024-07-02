import csv
import openpyxl
from fpdf import FPDF
import re
from typing import List, Dict, Callable


class Employee:
    emp_id: int = 1

    def __init__(self, name: str, email: str, department: str):
        """
        Initializes an Employee object with name, email, and department.
        """
        self.id: int = Employee.emp_id
        Employee.emp_id += 1
        self.name: str = name
        self.email: str = email
        self.department: str = department

    def __str__(self) -> str:
        """
        Returns a string representation of the employee.
        """
        return f"ID: {self.id}, Name: {self.name}, Email: {self.email}, Department: {self.department}"

    def validate(self) -> bool:
        """
        Validates the employee's email.
        Returns True if valid, False otherwise.
        """
        if not re.match(r"[^@]+@[^@]+\.[^@]+", self.email):
            return False
        return True

    def get_details(self) -> List[str]:
        """
        Returns the employee details as a list.
        """
        return [self.name, self.email, self.department]


class EmployeeList:
    def __init__(self):
        """
        Initializes an empty list of employees.
        """
        self.employees: List[Employee] = []

    def add_employee(self, emp: Employee) -> None:
        """
        Adds an employee to the list.
        """
        self.employees.append(emp)

    def remove_employee(self, emp_id: int) -> None:
        """
        Removes an employee from the list by ID.
        """
        self.employees = [emp for emp in self.employees if emp.id != emp_id]

    def display_employees(self) -> None:
        """
        Displays all employees in the list.
        """
        if not self.employees:
            print("No employees to display.")
        else:
            for emp in self.employees:
                print(emp)

    def save_to_csv(self, file_name: str) -> None:
        """
        Saves employee details to a CSV file.
        """
        with open(file_name, "w", newline='') as file:
            writer = csv.writer(file)
            writer.writerow(["Name", "Email", "Department"])
            for emp in self.employees:
                writer.writerow(emp.get_details())

    def save_to_excel(self, file_name: str) -> None:
        """
        Saves employee details to an Excel file.
        """
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["Name", "Email", "Department"])
        for emp in self.employees:
            sheet.append(emp.get_details())
        wb.save(file_name)

    def save_to_pdf(self, file_name: str) -> None:
        """
        Saves employee details to a PDF file.
        """
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        pdf.cell(200, 10, txt="Employee Details", ln=True, align="C")
        pdf.cell(40, 10, txt="Name", border=1)
        pdf.cell(70, 10, txt="Email", border=1)
        pdf.cell(40, 10, txt="Department", border=1)
        pdf.ln(10)

        for emp in self.employees:
            pdf.cell(40, 10, txt=emp.name, border=1)
            pdf.cell(70, 10, txt=emp.email, border=1)
            pdf.cell(40, 10, txt=emp.department, border=1)
            pdf.ln(10)

        pdf.output(file_name)

    def read_from_csv(self, file_name: str) -> None:
        """
        Reads employee details from a CSV file and adds them to the list.
        """
        with open(file_name, "r") as file:
            reader = csv.reader(file)
            next(reader)  # Skip header row
            for row in reader:
                if row:
                    name, email, department = row
                    emp = Employee(name, email, department)
                    self.add_employee(emp)

    def read_from_excel(self, file_name: str) -> None:
        """
        Reads employee details from an Excel file and adds them to the list.
        """
        wb = openpyxl.load_workbook(file_name)
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row:
                name, email, department = row
                emp = Employee(name, email, department)
                self.add_employee(emp)


def add_employee(emp_list: EmployeeList) -> None:
    """
    Adds a new employee to the employee list.
    """
    name = input("Enter name: ")
    email = input("Enter email: ")
    department = input("Enter department: ")
    emp = Employee(name, email, department)
    if emp.validate():
        emp_list.add_employee(emp)
        print("Employee added successfully")
    else:
        print("Invalid email address")


def display_employees(emp_list: EmployeeList) -> None:
    """
    Displays all employees in the employee list.
    """
    emp_list.display_employees()


def remove_employee(emp_list: EmployeeList) -> None:
    """
    Removes an employee from the employee list by ID.
    """
    try:
        emp_id = int(input("Enter Employee ID to remove: "))
        emp_list.remove_employee(emp_id)
        print(f"Employee with ID {emp_id} removed successfully")
    except ValueError:
        print("Invalid input. Please enter a number.")


def save_to_csv(emp_list: EmployeeList) -> None:
    """
    Saves the employee list to a CSV file.
    """
    file_name = input("Enter CSV file name: ")
    emp_list.save_to_csv(file_name)


def save_to_excel(emp_list: EmployeeList) -> None:
    """
    Saves the employee list to an Excel file.
    """
    file_name = input("Enter Excel file name: ")
    emp_list.save_to_excel(file_name)


def save_to_pdf(emp_list: EmployeeList) -> None:
    """
    Saves the employee list to a PDF file.
    """
    file_name = input("Enter PDF file name: ")
    emp_list.save_to_pdf(file_name)


def read_from_csv(emp_list: EmployeeList) -> None:
    """
    Reads employee details from a CSV file and adds them to the list.
    """
    file_name = input("Enter CSV file name: ")
    emp_list.read_from_csv(file_name)


def read_from_excel(emp_list: EmployeeList) -> None:
    """
    Reads employee details from an Excel file and adds them to the list.
    """
    file_name = input("Enter Excel file name: ")
    emp_list.read_from_excel(file_name)


def main() -> None:
    """
    Main function to run the employee management application.
    """
    emp_list = EmployeeList()

    # Dictionary to map choices to functions
    menu_options: Dict[int, Callable[[EmployeeList], None]] = {
        1: add_employee,
        2: display_employees,
        3: remove_employee,
        4: save_to_csv,
        5: save_to_excel,
        6: save_to_pdf,
        7: read_from_csv,
        8: read_from_excel,
    }

    while True:
        # Menu options
        print("\n1. Add Employee")
        print("2. Display Employees")
        print("3. Remove Employee")
        print("4. Save to CSV file")
        print("5. Save to Excel file")
        print("6. Save to PDF file")
        print("7. Read from CSV file")
        print("8. Read from Excel file")
        print("9. Exit")

        # Get user input and handle the choice
        try:
            choice = int(input("Enter your choice: "))
        except ValueError:
            print("Invalid input. Please enter a number.")
            continue

        if choice in menu_options:
            # Call the corresponding function from the dictionary
            menu_options[choice](emp_list)
        elif choice == 9:
            print("Exiting the application. Goodbye!")
            break
        else:
            print("Invalid choice. Please select a valid option.")


if __name__ == "__main__":
    main()

